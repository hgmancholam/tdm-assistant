"""
excel_report.py — Generate a formatted Excel status report for a project.

Input (stdin JSON): full project context from get-project.ps1 + optional extras
  {
    "projectCode":  "ALPHA",
    "projectName":  "My Project",
    "reportDate":   "2026-06-17",
    "summary":      "Project is on track...",
    "status":       "green|yellow|red",
    "sprints":      [...],     # optional — velocity data
    "risks":        [...],     # optional — risk register
    "budget":       {...},     # optional — budget info
    "actionItems":  [...]      # optional — action items
  }

Output: .xlsx file saved to projects/<CODE>/reports/
        prints JSON with file path to stdout
"""

import sys, json, argparse
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side, GradientFill)
from openpyxl.utils import get_column_letter


# ── Color palette ──────────────────────────────────────────────────────────────
C = {
    "dark_bg":    "1E1E2E",
    "header_bg":  "313244",
    "green":      "50FA7B",
    "yellow":     "F1FA8C",
    "red":        "FF5555",
    "blue":       "4E9AF1",
    "white":      "F8F8F2",
    "subheader":  "6272A4",
    "cell_alt":   "26263A",
}

STATUS_COLOR = {"green": C["green"], "yellow": C["yellow"], "red": C["red"]}

thin = Side(style="thin", color="444466")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _hdr(ws, row, col, value, bg=C["header_bg"], bold=True, size=11, color=C["white"]):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=bold, color=color, size=size, name="Calibri")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = border
    return cell


def _cell(ws, row, col, value, bold=False, color=C["white"], bg=C["dark_bg"], align="left"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=bold, color=color, size=10, name="Calibri")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border    = border
    return cell


def build_summary_sheet(ws, data):
    ws.sheet_properties.tabColor = STATUS_COLOR.get(data.get("status", "green"), C["green"])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60

    _hdr(ws, 1, 1, "PROJECT STATUS REPORT", bg=C["blue"], size=14)
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 30

    fields = [
        ("Project",      data.get("projectName", data.get("projectCode", ""))),
        ("Code",         data.get("projectCode", "")),
        ("Report Date",  data.get("reportDate", str(date.today()))),
        ("Status",       data.get("status", "").upper()),
        ("Phase",        data.get("phase", "")),
        ("Client",       data.get("client", "")),
    ]
    for i, (label, value) in enumerate(fields, start=2):
        _cell(ws, i, 1, label, bold=True, bg=C["header_bg"])
        color = STATUS_COLOR.get(data.get("status", ""), C["white"]) if label == "Status" else C["white"]
        _cell(ws, i, 2, value, color=color)

    row = len(fields) + 3
    _hdr(ws, row, 1, "EXECUTIVE SUMMARY", bg=C["header_bg"])
    ws.merge_cells(f"A{row}:B{row}")
    row += 1
    cell = ws.cell(row=row, column=1, value=data.get("summary", ""))
    cell.font      = Font(color=C["white"], size=10, name="Calibri")
    cell.fill      = PatternFill("solid", fgColor=C["dark_bg"])
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.border    = border
    ws.merge_cells(f"A{row}:B{row}")
    ws.row_dimensions[row].height = 80


def build_sprints_sheet(ws, sprints):
    ws.column_dimensions["A"].width = 18
    for col_letter in ["B", "C", "D", "E"]:
        ws.column_dimensions[col_letter].width = 16

    headers = ["Sprint", "Committed", "Completed", "Ratio %", "Status"]
    for c, h in enumerate(headers, start=1):
        _hdr(ws, 1, c, h)

    for r, sprint in enumerate(sprints, start=2):
        committed = sprint.get("committed", 0)
        completed = sprint.get("completed", 0)
        ratio     = round(completed / committed * 100, 1) if committed else 0
        status    = "green" if ratio >= 85 else "yellow" if ratio >= 70 else "red"
        bg        = C["cell_alt"] if r % 2 == 0 else C["dark_bg"]

        _cell(ws, r, 1, sprint.get("name", f"Sprint {r-1}"), bg=bg)
        _cell(ws, r, 2, committed, align="center", bg=bg)
        _cell(ws, r, 3, completed, align="center", bg=bg)
        _cell(ws, r, 4, f"{ratio}%", align="center", bg=bg)
        _cell(ws, r, 5, status.upper(), color=STATUS_COLOR[status], bold=True, align="center", bg=bg)


def build_risks_sheet(ws, risks):
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 35

    headers = ["Risk", "Probability", "Impact", "Score", "Mitigation"]
    for c, h in enumerate(headers, start=1):
        _hdr(ws, 1, c, h)

    for r, risk in enumerate(risks, start=2):
        prob   = risk.get("probability", 0)
        impact = risk.get("impact", 0)
        score  = prob * impact
        color  = C["red"] if score >= 15 else C["yellow"] if score >= 8 else C["green"]
        bg     = C["cell_alt"] if r % 2 == 0 else C["dark_bg"]

        _cell(ws, r, 1, risk.get("title", ""),                        bg=bg)
        _cell(ws, r, 2, f"{prob}/5",  align="center",                 bg=bg)
        _cell(ws, r, 3, f"{impact}/5", align="center",                bg=bg)
        _cell(ws, r, 4, score, align="center", color=color, bold=True, bg=bg)
        _cell(ws, r, 5, risk.get("mitigation", ""),                   bg=bg)


def build_actions_sheet(ws, action_items):
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12

    headers = ["Action Item", "Owner", "Due Date", "Status"]
    for c, h in enumerate(headers, start=1):
        _hdr(ws, 1, c, h)

    status_colors = {"done": C["green"], "in progress": C["yellow"], "pending": C["red"]}

    for r, item in enumerate(action_items, start=2):
        status = item.get("status", "pending").lower()
        color  = status_colors.get(status, C["white"])
        bg     = C["cell_alt"] if r % 2 == 0 else C["dark_bg"]

        _cell(ws, r, 1, item.get("text", ""),                          bg=bg)
        _cell(ws, r, 2, item.get("owner", ""),                         bg=bg)
        _cell(ws, r, 3, item.get("dueDate", ""),   align="center",     bg=bg)
        _cell(ws, r, 4, status.upper(), color=color, bold=True,
              align="center",                                            bg=bg)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", default=None)
    args = parser.parse_args()

    raw  = sys.stdin.read().strip()
    data = json.loads(raw) if raw else {}

    code       = data.get("projectCode", "PROJECT")
    output_dir = Path(args.output_dir) if args.output_dir else Path("projects") / code / "reports"
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    ws_summary = wb.create_sheet("Summary")
    build_summary_sheet(ws_summary, data)

    if data.get("sprints"):
        ws_sprints = wb.create_sheet("Sprint Velocity")
        build_sprints_sheet(ws_sprints, data["sprints"])

    if data.get("risks"):
        ws_risks = wb.create_sheet("Risk Register")
        build_risks_sheet(ws_risks, data["risks"])

    if data.get("actionItems"):
        ws_actions = wb.create_sheet("Action Items")
        build_actions_sheet(ws_actions, data["actionItems"])

    filename = output_dir / f"status-report-{code}-{date.today().isoformat()}.xlsx"
    wb.save(filename)

    print(json.dumps({
        "success":    True,
        "reportPath": str(filename),
        "sheets":     wb.sheetnames
    }, indent=2))


if __name__ == "__main__":
    main()
